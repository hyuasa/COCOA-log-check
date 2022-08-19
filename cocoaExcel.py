#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Excel関連モジュール

- シートの整形
- グラフ追加
- DataFrame形式をExcelにする
- Excel Bookを保管する

"""
import json  # json
import logging
import os  # ファイルシステム系　path など
import re  # 正規表現チェック
import sys  # process関係
import time  # sleep
from datetime import date, datetime, timedelta
from datetime import timezone as dttz  # date time 関係
from pprint import pformat, pprint  # format dump for List, Dictオブジェクト

import dateutil.parser  # date time 関係 python-dateutil
import openpyxl
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.formatting.rule import Rule
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.chart import Reference, BarChart, Series

import cocoaConfig as cc

__author__ = "hyuasa"
__version__ = "0.0.1"
__date__ = "Aug 16 2022"

# シート内最終グラフ位置
CURRENT_GRAPH_POSITION = 0

# color constant see https://www.colordic.org/w
COLOR_SHINBASHI = 'FFbce2e8'
COLOR_KONPEKI = 'FF007bbb'
COLOR_GEPPAKU = 'FFEAF4FC'
COLOR_BLACK = 'FFFFFFFF'
COLOR_KONAI = 'FF0f2350'
COLOR_AINEZU = 'FF6c848d'
BAR_COLOR_MIZUIRO = 'bce2e8'

# set font IBM Plex Suns JP エクストラ・ライト 10 point
NORMAL_FONT = Font(name=cc.FONT_NAME, size='10')
# LIGHT_FONT = Font(name=cc.FONT_NAME,size='9',color=COLOR_GEPPAKU)
DARK_FONT = Font(name=cc.FONT_NAME, size='9', color=COLOR_KONAI)
GRAY_FONT = Font(name=cc.FONT_NAME, size='9', color=COLOR_AINEZU)
HEADLINE_FONT = Font(name=cc.FONT_NAME, size='12', bold=True,
                     italic=True, shadow=True, color=COLOR_KONPEKI)
HEADLINE2_FONT = Font(name=cc.FONT_NAME, size='10', bold=True,
                      italic=True, shadow=True, color=COLOR_KONPEKI)
TITLE_CELL = PatternFill(patternType='solid', fgColor=COLOR_GEPPAKU)
TIME_CELL = PatternFill(patternType='solid', fgColor=COLOR_SHINBASHI)
# PATTERN = PatternFill( fill_type = None ,start_color = COLOR_SHINBASHI , end_color = COLOR_GEPPAKU )
AL_CENTER = Alignment(horizontal='center', vertical='center',
                      wrap_text=False, shrink_to_fit=False)
AL_TOPLEFT_WRAP = Alignment(horizontal='left', vertical='top',
                            wrap_text=True, shrink_to_fit=False)
AL_TOPLEFT = Alignment(horizontal='left', vertical='top',
                       wrap_text=False, shrink_to_fit=False)
AL_TOPCENTER = Alignment(horizontal='center', vertical='top',
                         wrap_text=False, shrink_to_fit=False)
RED_FONT = Font(name=cc.FONT_NAME, size='9', bold=True, color='FF0000')

NORMAL_BORDER = Border(left=Side(border_style=None, color=COLOR_BLACK),
                       right=Side(border_style=None, color=COLOR_BLACK),
                       top=Side(border_style=None, color=COLOR_BLACK),
                       bottom=Side(border_style=None, color=COLOR_BLACK),
                       diagonal=Side(border_style=None, color=COLOR_BLACK), diagonal_direction=0,
                       outline=Side(border_style=None, color=COLOR_BLACK),
                       vertical=Side(border_style=None, color=COLOR_BLACK),
                       horizontal=Side(border_style=None, color=COLOR_BLACK)
                       )


def stitle(ws, position):
    """指定された行をタイトル行として、カラム名と各種属性をリスト形式で返す

    Args:
        ws (Worksheet): openpyxl ワークシートオブジェクト
        position (int): タイトル行の位置　行番号

    Returns:
        (list): 以下例にあるdictのリスト

    Example:
        [{'name': カラム名, 'letter': 'A','column': 1, 'coordinate': 'A1'},{},{}]

    """
    # scan and return titles attribute at row position
    titles = []
    title_row = ws[position]
    for col in title_row:
        title = {'name': col.value, 'letter': col.column_letter,
                 'column': col.column, 'coordinate': col.coordinate}
        titles.append(title)
    return titles


def ftitle(titles, name):
    """stitleでスキャンしたtitlesのカラム名からカラム属性を返す

    Args:
        titles (list): stitleでスキャンしたカラムのリスト
        name (str): カラム名

    Returns:
        (dict): 以下の例にあるカラム属性

    Example:
        [{'name': カラム名, 'letter': 'A','column': 1, 'coordinate': 'A1'},{},{}]
    """
    # find and return name of title attribute
    title = [title for title in titles if title['name'] == name]
    if len(title) >= 1:
        return title[0]
    else:
        # 見つけられなかった時はエラーを防ぐためにA1を戻す
        return {'name': 'unknown', 'letter': 'A', 'column': 1, 'coordinate': 'A1'}


def add_title_comment(logger, wb, sheetname, titlerow, titlename, comment):
    """ワークブックのシート名タイトル行、カラム名とそこに追加するコメント文字列を渡してセルにコメントを付ける

    Args:
        logger (logger): ロギングオブジェクト
        wb (Workbook): Excel workbookオブジェクト
        sheetname (str): ワークシート名
        titlerow (int): タイトルの行番号
        titlename (str): タイトルカラムの名前
        comment (str): 追加するコメントストリング

    Returns:
        (Workbook): Workbookオブジェクト 保管されていない
    """
    logger.info(f'adding comment to {titlename} in sheet {sheetname}')
    ws = wb[sheetname]
    titles = stitle(ws, titlerow)
    title = ftitle(titles, titlename)
    comment.width = 500
    comment.height = 300
    ws[title['coordinate']].comment = comment
    return wb


def add_comment_direct(logger, wb, sheetname, coordinate, comment):
    """Excelのワークブック、シート名、座標と追加するコメントを受けて、その位置にコメントを追加する

    Args:
        logger (logger): ロギングオブジェクト
        wb (Workbook): Excel workbookオブジェクト
        sheetname (str): ワークシート名
        coordinate (str): ワークシート上の座標　ex: 'A1'
        comment (str): 追加するコメントストリング

    Returns:
        (Workbook): Workbookオブジェクト 保管されていない
    """
    logger.info('adding comment to {} in sheet {}'.format(
        coordinate, sheetname))
    ws = wb[sheetname]
    comment.width = 500
    comment.height = 300
    ws[coordinate].comment = comment
    return wb


def shape_exposure_sheet(logger, ws):
    """接触履歴ワークシート整形関数

    Args:
        logger (logger): ロギングオブジェクト
        ws (Wroksheet): ワークシートオブジェクト

    Return:
        None
    """
    maxrow = ws.max_row
    maxcolumn = ws.max_column
    logger.info(f'maxcolumn: {maxcolumn} maxrow:{maxrow}')
    ws.sheet_properties.tabColor = '007bbb'     # 紺碧

    titles3 = stitle(ws, 3)  # 3行目ラベル
    titles4 = stitle(ws, 4)  # 4行目ラベル
    ws.column_dimensions[ftitle(titles4, 'date')['letter']].width = 15
    ws.column_dimensions[ftitle(titles3, '接触時間計(分)')['letter']].width = 12
    ws.column_dimensions[ftitle(titles3, 'contact')['letter']].width = 12
    ws.column_dimensions[ftitle(titles3, 'cocoa_score')['letter']].width = 12
    ws.column_dimensions[ftitle(titles3, '算出スコア計')['letter']].width = 12
    ws['B3'].value = '距離 >>'
    ws['A4'].value = '接触日'
    ws['B4'].value = ''   # dow 曜日ラベル

    # 閾値越えスコアの赤字表示
    dxf = DifferentialStyle(font=RED_FONT)
    rule = Rule(type='cellIs', operator='greaterThanOrEqual',
                formula=[cc.COCOA_SCORE_THRESHOLD], dxf=dxf)
    range_cocoa_score = ftitle(titles3, 'cocoa_score')['letter']+str(4)+':' + \
        ftitle(titles3, 'cocoa_score')['letter']+str(maxrow)
    ws.conditional_formatting.add(range_cocoa_score, rule)
    range_calculate_score = ftitle(titles3, '算出スコア計')['letter']+str(4)+':' + \
        ftitle(titles3, '算出スコア計')['letter']+str(maxrow)
    ws.conditional_formatting.add(range_calculate_score, rule)

    range_contact_duration = ftitle(titles3, '接触時間計(分)')['letter']+str(4)+':' + \
        ftitle(titles3, '接触時間計(分)')['letter']+str(maxrow)

    fill_cell_color(ws[range_cocoa_score], COLOR_GEPPAKU)
    fill_cell_color(ws[range_calculate_score], COLOR_GEPPAKU)
    fill_cell_color(ws[range_contact_duration], COLOR_GEPPAKU)

    return


def fill_cell_color(cell_range, color):
    """セル色設定

    Args:
        cell_range (Range): ワークシート行レンジオブジェクト ws[A1:A4]
        color (str): 色

    Returns:
        None
    """
    fill = PatternFill(patternType='solid', fgColor=color)
    for row in cell_range:
        for cell in row:
            cell.fill = fill
    return


def shape_sheet_common(logger, ws):
    """ワークシート共通で使用する整形関数

    - Font設定

    Args:
        logger (logger): ロギングオブジェクト
        ws (Wroksheet): ワークシートオブジェクト

    Returns:
        None
    """
    logger.info(f'set font to columns in the sheet: {ws.title}')
    maxrow = ws.max_row
    maxcolumn = ws.max_column

    for row in ws:
        for cell in row:
            ws[cell.coordinate].font = NORMAL_FONT
        for c in range(1, maxcolumn+1):
            ws[ws.cell(row=1, column=c).coordinate].fill = TITLE_CELL
            ws[ws.cell(row=1, column=c).coordinate].border = NORMAL_BORDER

    return


def save_book(logger, wb, bookname):
    """Excel ワークブックを保管する

    Caution:
        大きなBookの保管は時間がかかるので、この関数の呼び出しを最小限にすることでパフォーマンス向上が期待できる

    Args:
        logger (logger): ロギングオブジェクト
        wb (Workbook): openpyxlのワークブックオブジェクト
        bookname (str): Excelファイル名

    Returns:
        None
    """
    logger.info(f'saving book: {bookname}')
    wb.save(bookname)
    logger.info(f'book saved : {bookname}')
    return


def save_to_excel_multi(logger, bookname=None, dfs=None, sheets=None, indexes=None):
    """複数のデータフレームをExcelに抽出する

    Args:
        logger (_type_): _description_
        bookname (str): Excelファイル名
        dfs (list of dataframe): 抽出するデータフレームのリスト
        sheets (list of str): 抽出先のシート名のリスト
        indexes (list of boolian ): indexを含めて抽出するかどうかをデータフレーム毎にリストで

    Returns:
        (Workbook) : Workbook object
    """
    logger.info('export to book: {}'.format(bookname))
    with pd.ExcelWriter(bookname) as writer:
        for i in range(len(dfs)):
            dfs[i].to_excel(writer, sheet_name=sheets[i], index=indexes[i])
    logger.info('get wb object of book: {}'.format(bookname))
    wb = openpyxl.load_workbook(bookname)
    logger.info('returns wb object')
    return wb


def shape_a_sheets(logger, wb):
    """シート整形のメイン

    Args:
        logger (logger): ロギングオブジェクト
        wb (Workbook): Workbookオブジェクト

    Returns:
        (Workbook): Workbookオブジェクト
    """
    sheetnames = wb.sheetnames
    for i in range(len(sheetnames)):
        ws = wb.worksheets[i]

        if ws.title == cc.COCOA_EXPOSURE_SHEET_NAME:
            shape_sheet_common(logger, ws)
            shape_exposure_sheet(logger, ws)
        else:
            shape_sheet_common(logger, ws)

    return wb


def add_chart(logger, wb, element, row=3, ctitle='', x_title='日付', y_title=''):
    """COCOA各種グラフの追加
       要素(element)１つの棒グラフ

    Args:
        logger (logger): ロギングオブジェクト
        wb (Workbook): Workbookオブジェクト
        element(str): グラフ要素のカラム名
        row(int): 要素名のある行 default 3, 
        ctitle(str): チャートタイトル
        x_title(str): X軸タイトル defult:'日付'
        y_title(str): Y軸タイトル

    Returns:
        (Workbook): Workbookオブジェクト
    """
    global CURRENT_GRAPH_POSITION

    ws = wb[cc.COCOA_EXPOSURE_SHEET_NAME]
    maxrow = ws.max_row
    titles3 = stitle(ws, row)  # defult 3行目ラベル
    min_col = ftitle(titles3, element)['column']
    max_col = min_col

    chart = BarChart()
    src = Reference(ws, min_col=min_col, min_row=5,
                    max_col=max_col, max_row=maxrow)
    cat = Reference(ws, min_col=1, min_row=5, max_row=maxrow,)
    chart.add_data(src, titles_from_data=False)
    chart.set_categories(cat)
    chart.legend = None
    chart.gapWidth = 10
    if CURRENT_GRAPH_POSITION == 0:
        CURRENT_GRAPH_POSITION = maxrow+2
    chart.anchor = 'A' + str(CURRENT_GRAPH_POSITION)
    CURRENT_GRAPH_POSITION += 22
    chart.title = ctitle
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.width = 30
    chart.height = 10
    series1 = chart.series[0]
    series1.spPr.solidFill = BAR_COLOR_MIZUIRO
    ws.add_chart(chart)

    return wb


def create_cocoa_excel(logger, merge_df):
    """create cocoa log Excel book

    Args:
        logger (logging): ロガー
        merge_df (DataFrame): マージ後のDataFrame

    Returns:
        None

    """

    # Excel保管
    bookname = 'COCOA_LOG_CHECKER_' + \
        datetime.now(cc.JST).strftime('%Y-%m-%d-%H%M')+'.xlsx'
    wb = save_to_excel_multi(logger, bookname=bookname,
                             dfs=[merge_df],
                             sheets=[cc.COCOA_EXPOSURE_SHEET_NAME],
                             indexes=[True])
    #　Excelシート整形
    wb = shape_a_sheets(logger, wb)
    comment = Comment('スコア1350以上が濃厚接触アラート対象になるようです', 'cocoa_log_checker')
    wb = add_title_comment(
        logger, wb, cc.COCOA_EXPOSURE_SHEET_NAME, 3, 'cocoa_score', comment)
    wb = add_title_comment(
        logger, wb, cc.COCOA_EXPOSURE_SHEET_NAME, 3, '算出スコア計', comment)
    wb = add_chart(logger, wb, 'cocoa_score',
                   ctitle='COCOA Score', y_title='スコア')
    wb = add_chart(logger, wb, '算出スコア計',
                   ctitle='COCOA Calculated Score', y_title='スコア')
    wb = add_chart(logger, wb, 'contact', ctitle='接触回数', y_title='回数')
    wb = add_chart(logger, wb, '接触時間計(分)', ctitle='接触時間(分)', y_title='分')
    save_book(logger, wb, bookname)

    return
